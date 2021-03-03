#!/usr/bin/env python

import sys
import csv
from pathlib import Path
from openpyxl import load_workbook

def save_csv(rows, fp_out):
    fp_out = Path(fp_out)
    fp_out.parent.mkdir(parents=True, exist_ok=True)
    with open(fp_out, "wt") as f_out:
        writer = csv.DictWriter(f_out, fieldnames=rows[0].keys(), lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)

def convert_s5(fp_in, outdir="parsed"):
    outdir = Path(outdir)
    wb = load_workbook(filename = fp_in)
    s5a = parse_s5_ighv(wb.worksheets[0])
    save_csv(s5a, outdir / "S5A.csv")
    s5b = parse_s5_ighv(wb.worksheets[1], 3)
    save_csv(s5b, outdir / "S5B.csv")
    s5c = parse_s5_ighv(wb.worksheets[2], 3)
    save_csv(s5c, outdir / "S5C.csv")
    s5d = _parse(wb.worksheets[3], 2, 1, 6, 3)
    save_csv(s5d, outdir / "S5D.csv")
    s5e = _parse(wb.worksheets[4], 2, 1, 6, 3)
    save_csv(s5e, outdir / "S5E.csv")

def parse_s5_ighv(sheet, rowstart=2):
    offsets = []
    colidx = 1
    while True:
        if not sheet.cell(rowstart, colidx).value:
            break
        if sheet.cell(rowstart, colidx).value == "Alleles":
            offsets.append(colidx)
        colidx += 1
    rows = []
    for offset in offsets:
        rows += _parse(sheet, rowstart, offset, 6, 6)
    return rows

def _parse(sheet, rowstart, colstart, cols_text, cols_color):
    rows = []
    headings = [sheet.cell(rowstart, colstart + x).value for x in range(cols_text + cols_color)]
    # loop over rows until we're past the bottom
    rowidx = 1
    while True:
        row = {key: "" for key in headings}
        for colidx in range(cols_text):
            row[headings[colidx]] = sheet.cell(rowstart+rowidx, colstart + colidx).value
        for colidx in range(cols_color):
            pattern_types = {
                "solid": "T",
                None: "F"}
            pattern_type = sheet.cell(rowstart+rowidx, colstart + colidx + cols_text).fill.patternType
            try:
                row[headings[colidx+cols_text]] = pattern_types[pattern_type]
            except KeyError as err:
                print("cell %d, %d" % (rowstart+rowidx, colstart+colidx+cols_text))
                raise err
        if not row["Alleles"]:
            break
        rowidx += 1
        rows.append(row)
    return rows

if __name__ == "__main__":
    convert_s5(sys.argv[1])
